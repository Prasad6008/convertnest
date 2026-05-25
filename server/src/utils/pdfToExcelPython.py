import sys
import re
import statistics
import fitz
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


HEADERS = [
    "Country",
    "Year",
    "Rank",
    "Total",
    "S1: Demographic Pressures",
    "S2: Refugees and IDPs",
    "C3: Group Grievance",
    "E3: Human Flight and Brain Drain",
    "E2: Economic Inequality",
    "E1: Economy",
    "P1: State Legitimacy",
    "P2: Public Services",
    "P3: Human Rights",
    "C1: Security Apparatus",
    "C2: Factionalized Elites",
    "X1: External Intervention"
]

YEAR_RE = re.compile(r"^20\d{2}$")
RANK_RE = re.compile(r"^\d+(st|nd|rd|th)$", re.IGNORECASE)
NUMBER_RE = re.compile(r"^\d+(\.\d+)?$")

EXPECTED_NUMERIC_VALUES = 13


def clean_text(value):
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()


def is_year(value):
    return bool(YEAR_RE.match(clean_text(value)))


def is_rank(value):
    return bool(RANK_RE.match(clean_text(value)))


def is_number(value):
    return bool(NUMBER_RE.match(clean_text(value)))


def y_center(word):
    return (word[1] + word[3]) / 2


def x_center(word):
    return (word[0] + word[2]) / 2


def get_clean_words(page):
    words = []
    for word in page.get_text("words"):
        text = clean_text(word[4])
        if text:
            words.append(word)
    return words


def cluster_values(values, tolerance=5):
    if not values:
        return []

    values = sorted(values)
    clusters = []

    for value in values:
        if not clusters:
            clusters.append([value])
            continue

        avg = sum(clusters[-1]) / len(clusters[-1])

        if abs(value - avg) <= tolerance:
            clusters[-1].append(value)
        else:
            clusters.append([value])

    return [sum(cluster) / len(cluster) for cluster in clusters]


def find_row_anchors(words):
    year_positions = [
        y_center(word)
        for word in words
        if is_year(clean_text(word[4]))
    ]

    anchors = cluster_values(year_positions, tolerance=5)
    valid = []

    for anchor in anchors:
        nearby = [
            word for word in words
            if abs(y_center(word) - anchor) <= 9
        ]

        texts = [clean_text(word[4]) for word in nearby]

        has_year = any(is_year(text) for text in texts)
        has_rank = any(is_rank(text) for text in texts)
        numeric_count = sum(
            1 for text in texts
            if is_number(text) and not is_year(text)
        )

        if has_year and has_rank and numeric_count >= 3:
            valid.append(anchor)

    return sorted(valid)


def build_row_bands(anchors):
    if not anchors:
        return []

    bands = []

    for index, anchor in enumerate(anchors):
        if index == 0:
            gap = anchors[1] - anchors[0] if len(anchors) > 1 else 18
            top = anchor - gap / 2
        else:
            top = (anchors[index - 1] + anchor) / 2

        if index == len(anchors) - 1:
            gap = anchors[-1] - anchors[-2] if len(anchors) > 1 else 18
            bottom = anchor + gap / 2
        else:
            bottom = (anchor + anchors[index + 1]) / 2

        bands.append({
            "top": top,
            "bottom": bottom,
            "anchor": anchor
        })

    return bands


def words_in_band(words, top, bottom):
    return sorted(
        [word for word in words if top <= y_center(word) < bottom],
        key=lambda word: word[0]
    )


def find_first_word(row_words, condition, after_x=None):
    for word in row_words:
        text = clean_text(word[4])

        if after_x is not None and word[0] <= after_x:
            continue

        if condition(text):
            return word

    return None


def extract_country(row_words, year_word):
    country_words = []

    for word in row_words:
        if word[2] < year_word[0] - 2:
            country_words.append(clean_text(word[4]))

    country = " ".join(country_words).strip()
    country = re.sub(r"\s+", " ", country)

    return country


def median_gap(anchors):
    if len(anchors) < 2:
        return 18

    gaps = [
        anchors[index + 1] - anchors[index]
        for index in range(len(anchors) - 1)
    ]

    return statistics.median(gaps)


def nearest_anchor(y, anchors, max_distance):
    if not anchors:
        return None

    nearest = min(anchors, key=lambda anchor: abs(anchor - y))

    if abs(nearest - y) <= max_distance:
        return nearest

    return None


def collect_numeric_items(words, anchors, rank_x_min):
    gap = median_gap(anchors)
    max_distance = max(10, min(24, gap * 0.75))

    items = []

    for word in words:
        text = clean_text(word[4])

        if not is_number(text) or is_year(text):
            continue

        if word[0] <= rank_x_min:
            continue

        assigned_y = nearest_anchor(y_center(word), anchors, max_distance)

        if assigned_y is None:
            continue

        items.append({
            "row_y": assigned_y,
            "x": x_center(word),
            "value": text
        })

    return items


def cluster_numeric_columns(items):
    if not items:
        return []

    sorted_items = sorted(items, key=lambda item: item["x"])
    clusters = []

    for item in sorted_items:
        if not clusters:
            clusters.append([item])
            continue

        avg_x = sum(i["x"] for i in clusters[-1]) / len(clusters[-1])

        if abs(item["x"] - avg_x) <= 18:
            clusters[-1].append(item)
        else:
            clusters.append([item])

    columns = []

    for cluster in clusters:
        avg_x = sum(item["x"] for item in cluster) / len(cluster)
        columns.append({
            "x": avg_x,
            "count": len(cluster)
        })

    if len(columns) > EXPECTED_NUMERIC_VALUES:
        columns = sorted(columns, key=lambda col: col["count"], reverse=True)
        columns = columns[:EXPECTED_NUMERIC_VALUES]

    columns = sorted(columns, key=lambda col: col["x"])

    return [col["x"] for col in columns]


def column_boundaries(columns):
    boundaries = [-999999]

    for index in range(len(columns) - 1):
        boundaries.append((columns[index] + columns[index + 1]) / 2)

    boundaries.append(999999)

    return boundaries


def get_column_index(x, boundaries):
    for index in range(len(boundaries) - 1):
        if boundaries[index] <= x < boundaries[index + 1]:
            return index
    return None


def build_values_for_row(row_y, numeric_items, numeric_columns):
    values = [""] * EXPECTED_NUMERIC_VALUES

    boundaries = column_boundaries(numeric_columns)

    row_items = [
        item for item in numeric_items
        if abs(item["row_y"] - row_y) <= 0.1
    ]

    for item in row_items:
        col_index = get_column_index(item["x"], boundaries)

        if col_index is None:
            continue

        if 0 <= col_index < EXPECTED_NUMERIC_VALUES:
            existing = values[col_index]

            if not existing:
                values[col_index] = item["value"]
            else:
                # If duplicate value detected in same column, keep the one closer to column center
                current_distance = abs(item["x"] - numeric_columns[col_index])
                old_items = [
                    old for old in row_items
                    if old["value"] == existing
                ]

                if old_items:
                    old_distance = abs(old_items[0]["x"] - numeric_columns[col_index])
                    if current_distance < old_distance:
                        values[col_index] = item["value"]

    return values


def parse_row(row_words, row_y, numeric_items, numeric_columns):
    if not row_words:
        return None

    year_word = find_first_word(row_words, is_year)

    if year_word is None:
        return None

    rank_word = find_first_word(row_words, is_rank, after_x=year_word[2])

    if rank_word is None:
        return None

    country = extract_country(row_words, year_word)
    year = clean_text(year_word[4])
    rank = clean_text(rank_word[4])

    if not country:
        return None

    values = build_values_for_row(row_y, numeric_items, numeric_columns)

    row = [country, year, rank] + values

    if len(row) < len(HEADERS):
        row += [""] * (len(HEADERS) - len(row))

    return row[:len(HEADERS)]


def convert_numbers(ws, start_row, end_row, start_col, end_col):
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            value = cell.value

            if isinstance(value, str) and is_number(value):
                try:
                    if "." in value:
                        cell.value = float(value)
                    else:
                        cell.value = int(value)
                except Exception:
                    pass


def apply_style(ws, total_rows, total_cols):
    header_fill = PatternFill("solid", fgColor="EAF3F8")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    body_font = Font(name="Courier New", size=11)
    header_font = Font(name="Courier New", size=11, bold=True)

    ws.freeze_panes = "A2"

    for row in ws.iter_rows(min_row=1, max_row=total_rows, min_col=1, max_col=total_cols):
        for cell in row:
            cell.border = border
            cell.font = body_font
            cell.alignment = Alignment(
                vertical="center",
                horizontal="center",
                wrap_text=True
            )

    for col in range(1, total_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill

        if col <= 4:
            cell.alignment = Alignment(
                horizontal="center",
                vertical="bottom",
                wrap_text=True
            )
        else:
            cell.alignment = Alignment(
                textRotation=90,
                horizontal="center",
                vertical="bottom",
                wrap_text=True
            )

    for row in range(2, total_rows + 1):
        ws.cell(row=row, column=1).alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=False
        )

    ws.row_dimensions[1].height = 150

    for row in range(2, total_rows + 1):
        ws.row_dimensions[row].height = 18

    for col in range(1, total_cols + 1):
        letter = get_column_letter(col)

        if col == 1:
            ws.column_dimensions[letter].width = 18
        elif col in [2, 3, 4]:
            ws.column_dimensions[letter].width = 12
        else:
            ws.column_dimensions[letter].width = 12

    last_col = get_column_letter(total_cols)
    ws.auto_filter.ref = f"A1:{last_col}{total_rows}"


def pdf_to_excel(input_pdf, output_xlsx):
    pdf = fitz.open(input_pdf)
    all_rows = []

    for page in pdf:
        words = get_clean_words(page)

        if not words:
            continue

        anchors = find_row_anchors(words)
        bands = build_row_bands(anchors)

        if not anchors:
            continue

        # Find rank x position safely
        rank_x_values = []

        for band in bands:
            row_words = words_in_band(words, band["top"], band["bottom"])
            year_word = find_first_word(row_words, is_year)

            if year_word:
                rank_word = find_first_word(row_words, is_rank, after_x=year_word[2])
                if rank_word:
                    rank_x_values.append(rank_word[2])

        rank_x_min = min(rank_x_values) if rank_x_values else 0

        numeric_items = collect_numeric_items(words, anchors, rank_x_min)
        numeric_columns = cluster_numeric_columns(numeric_items)

        for band in bands:
            row_words = words_in_band(words, band["top"], band["bottom"])
            row = parse_row(
                row_words,
                band["anchor"],
                numeric_items,
                numeric_columns
            )

            if row:
                all_rows.append(row)

    pdf.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    total_cols = len(HEADERS)

    if not all_rows:
        ws.cell(row=1, column=1, value="No readable table found in this PDF.")
        wb.save(output_xlsx)
        return

    for col_index, header in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=col_index, value=header)

    for row_index, row in enumerate(all_rows, start=2):
        for col_index, value in enumerate(row, start=1):
            ws.cell(row=row_index, column=col_index, value=clean_text(value))

    total_rows = len(all_rows) + 1

    convert_numbers(
        ws,
        start_row=2,
        end_row=total_rows,
        start_col=4,
        end_col=total_cols
    )

    apply_style(ws, total_rows, total_cols)

    wb.save(output_xlsx)


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python pdfToExcelPython.py input.pdf output.xlsx", file=sys.stderr)
        sys.exit(1)

    input_pdf = sys.argv[1]
    output_xlsx = sys.argv[2]

    try:
        pdf_to_excel(input_pdf, output_xlsx)
        print("PDF to Excel conversion completed")
    except Exception as error:
        print(str(error), file=sys.stderr)
        sys.exit(1)